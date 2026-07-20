from __future__ import annotations

import argparse
import asyncio
import csv
import json
import os
import re
import sys
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse

from .dictionary import PROJECT_ROOT, normalize_text
from .youtube_api import (
    YouTubeApiError,
    YouTubeComment,
    YouTubeVideoApiClient,
    YouTubeVideoMetadata,
)
from .yt_search import SearchDependencyError


VIDEO_ID_PATTERN = re.compile(r"^[A-Za-z0-9_-]{11}$")
VIDEO_ID_HEADERS = {"videoid", "idvideo", "youtubeid", "idyoutube"}
VIDEO_URL_HEADERS = {"url", "videourl", "youtubeurl", "enlace", "link"}
INSTITUTION_HEADERS = {"institucion", "universidad", "institution"}
TITLE_HEADERS = {"titulo", "title", "videotitle"}
SOURCE_DATE_HEADERS = {"fecha", "date", "source_date", "fecharegistro"}


@dataclass(frozen=True)
class VideoRequest:
    video_id: str
    url: str
    row_number: int
    institution: str
    source_title: str | None = None
    source_date: str | None = None


@dataclass(frozen=True)
class DownloadResult:
    video_id: str
    url: str
    row_number: int
    institution: str
    status: str
    file_path: str | None = None
    error: str | None = None


@dataclass(frozen=True)
class CommentDownloadResult:
    video_id: str
    url: str
    row_number: int
    institution: str
    status: str
    comment_count: int
    file_path: str
    error: str | None = None


def _header_key(value: Any) -> str:
    return normalize_text(str(value or "")).replace("_", "").replace("-", "")


def _optional_cell_text(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if hasattr(value, "isoformat"):
        return str(value.isoformat())
    return str(value).strip() or None


def _video_id_from_value(value: Any) -> str | None:
    if not isinstance(value, str):
        return None
    candidate = value.strip()
    if VIDEO_ID_PATTERN.fullmatch(candidate):
        return candidate

    parsed = urlparse(candidate)
    if parsed.netloc.lower().endswith("youtu.be"):
        video_id = parsed.path.strip("/").split("/")[0]
    elif "youtube.com" in parsed.netloc.lower():
        video_id = parse_qs(parsed.query).get("v", [""])[0]
        if not video_id:
            path_parts = [part for part in parsed.path.split("/") if part]
            video_id = path_parts[-1] if path_parts and path_parts[0] in {"shorts", "embed"} else ""
    else:
        return None
    return video_id if VIDEO_ID_PATTERN.fullmatch(video_id) else None


def _institution_folder(institution: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", normalize_text(institution)).strip("_")
    return slug or "sin_institucion"


def _load_workbook_rows(input_path: Path, sheet_name: str | None) -> list[tuple[int, dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SearchDependencyError(
            "Falta openpyxl. Instala las dependencias con: pip install -r requirements.txt"
        ) from exc

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"La hoja {sheet_name!r} no existe. Opciones: {workbook.sheetnames}"
                )
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            raise ValueError("El archivo XLSX no contiene encabezados.")

        indexed_headers = {
            _header_key(header): index
            for index, header in enumerate(headers)
            if header not in (None, "")
        }
        video_id_index = next(
            (indexed_headers[header] for header in VIDEO_ID_HEADERS if header in indexed_headers),
            None,
        )
        video_url_index = next(
            (indexed_headers[header] for header in VIDEO_URL_HEADERS if header in indexed_headers),
            None,
        )
        institution_index = next(
            (indexed_headers[header] for header in INSTITUTION_HEADERS if header in indexed_headers),
            None,
        )
        title_index = next(
            (indexed_headers[header] for header in TITLE_HEADERS if header in indexed_headers),
            None,
        )
        source_date_index = next(
            (
                indexed_headers[header]
                for header in SOURCE_DATE_HEADERS
                if header in indexed_headers
            ),
            None,
        )
        if video_id_index is None and video_url_index is None:
            raise ValueError(
                "No se encontro una columna de video. Usa Video_id, video_id, url o youtube_url."
            )
        if institution_index is None:
            raise ValueError(
                "No se encontro una columna de institucion. Usa Institucion, universidad o institution."
            )

        def cell_value(row: tuple[Any, ...], index: int | None) -> Any:
            return row[index] if index is not None and index < len(row) else None

        loaded_rows: list[tuple[int, dict[str, Any]]] = []
        for row_number, row in enumerate(rows, start=2):
            loaded_rows.append(
                (
                    row_number,
                    {
                        "video_id": cell_value(row, video_id_index),
                        "url": cell_value(row, video_url_index),
                        "institution": cell_value(row, institution_index),
                        "title": cell_value(row, title_index),
                        "source_date": cell_value(row, source_date_index),
                    },
                )
            )
        return loaded_rows
    finally:
        workbook.close()


def load_video_requests(input_path: Path, sheet_name: str | None = None) -> list[VideoRequest]:
    if not input_path.is_file():
        raise ValueError(f"No existe el archivo de entrada: {input_path}")
    if input_path.suffix.lower() != ".xlsx":
        raise ValueError("El archivo de entrada debe tener extension .xlsx.")

    requests: list[VideoRequest] = []
    seen: set[tuple[str, str]] = set()
    for row_number, row in _load_workbook_rows(input_path, sheet_name):
        video_id = _video_id_from_value(row["video_id"]) or _video_id_from_value(row["url"])
        institution = str(row["institution"] or "").strip()
        deduplication_key = (video_id or "", _institution_folder(institution))
        if not video_id or deduplication_key in seen:
            continue
        seen.add(deduplication_key)
        requests.append(
            VideoRequest(
                video_id=video_id,
                url=f"https://www.youtube.com/watch?v={video_id}",
                row_number=row_number,
                institution=institution,
                source_title=_optional_cell_text(row["title"]),
                source_date=_optional_cell_text(row["source_date"]),
            )
        )
    if not requests:
        raise ValueError("No se encontraron IDs o URLs de YouTube validos en el XLSX.")
    return requests


def build_direct_video_requests(video_ids: list[str], institution: str) -> list[VideoRequest]:
    institution = institution.strip()
    if not institution:
        raise ValueError("Debes enviar --institution al usar --video-id.")

    requests: list[VideoRequest] = []
    seen: set[str] = set()
    for position, video_id in enumerate(video_ids, start=1):
        normalized_id = _video_id_from_value(video_id)
        if not normalized_id:
            raise ValueError(f"ID o URL de YouTube invalido: {video_id}")
        if normalized_id in seen:
            continue
        seen.add(normalized_id)
        requests.append(
            VideoRequest(
                video_id=normalized_id,
                url=f"https://www.youtube.com/watch?v={normalized_id}",
                row_number=position,
                institution=institution,
            )
        )
    return requests


def _download_one(
    request: VideoRequest,
    *,
    download_root: Path,
    video_format: str,
    retries: int,
    cookies_from_browser: str | None,
    cookies_browser_profile: str | None,
) -> DownloadResult:
    try:
        from yt_dlp import YoutubeDL
        from yt_dlp.utils import DownloadError
    except ImportError as exc:
        raise SearchDependencyError(
            "Falta yt-dlp. Instala las dependencias con: pip install -r requirements.txt"
        ) from exc

    output_dir = download_root / _institution_folder(request.institution)
    output_dir.mkdir(parents=True, exist_ok=True)
    options: dict[str, Any] = {
        "format": video_format,
        "outtmpl": str(output_dir / "%(id)s.%(ext)s"),
        "merge_output_format": "mp4",
        "noplaylist": True,
        "continuedl": True,
        "overwrites": False,
        "retries": retries,
        "fragment_retries": retries,
        "quiet": True,
        "no_warnings": True,
    }
    if cookies_from_browser:
        options["cookiesfrombrowser"] = (
            cookies_from_browser,
            cookies_browser_profile,
            None,
            None,
        )

    try:
        with YoutubeDL(options) as ydl:
            ydl.extract_info(request.url, download=True)
    except DownloadError as exc:
        return DownloadResult(
            video_id=request.video_id,
            url=request.url,
            row_number=request.row_number,
            institution=request.institution,
            status="failed",
            error=str(exc),
        )

    downloaded_files = sorted(
        path
        for path in output_dir.glob(f"{request.video_id}.*")
        if path.suffix not in {".part", ".ytdl", ".json"}
    )
    return DownloadResult(
        video_id=request.video_id,
        url=request.url,
        row_number=request.row_number,
        institution=request.institution,
        status="completed",
        file_path=str(downloaded_files[0]) if downloaded_files else None,
    )


async def download_videos(
    requests: list[VideoRequest],
    *,
    download_root: Path,
    workers: int = 1,
    video_format: str = "bv*[height<=720][ext=mp4]+ba[ext=m4a]/b[height<=720][ext=mp4]/b[ext=mp4]/b",
    retries: int = 0,
    cookies_from_browser: str | None = None,
    cookies_browser_profile: str | None = None,
) -> list[DownloadResult]:
    if workers < 1:
        raise ValueError("workers debe ser mayor que cero.")
    if retries < 0:
        raise ValueError("retries no puede ser negativo.")

    download_root.mkdir(parents=True, exist_ok=True)
    semaphore = asyncio.Semaphore(workers)

    async def run_one(request: VideoRequest) -> DownloadResult:
        async with semaphore:
            return await asyncio.to_thread(
                _download_one,
                request,
                download_root=download_root,
                video_format=video_format,
                retries=retries,
                cookies_from_browser=cookies_from_browser,
                cookies_browser_profile=cookies_browser_profile,
            )

    return await asyncio.gather(*(run_one(request) for request in requests))


COMMENT_FIELDS = [
    "record_status",
    "error",
    "video_id",
    "video_url",
    "video_title",
    "video_channel",
    "video_upload_date",
    "video_reported_comment_count",
    "institution",
    "source_row",
    "source_date",
    "comment_id",
    "parent_comment_id",
    "is_reply",
    "author_display_name",
    "author_channel_id",
    "comment_text",
    "like_count",
    "comment_published_at",
    "comment_updated_at",
    "reply_count",
]


def _comment_row(
    request: VideoRequest,
    metadata: YouTubeVideoMetadata,
    *,
    status: str,
    comment: YouTubeComment | None = None,
    error: str | None = None,
) -> dict[str, Any]:
    row: dict[str, Any] = {
        "record_status": status,
        "error": error,
        "video_id": request.video_id,
        "video_url": request.url,
        "video_title": metadata.title or request.source_title,
        "video_channel": metadata.channel_title,
        "video_upload_date": metadata.upload_date,
        "video_reported_comment_count": metadata.comment_count,
        "institution": request.institution,
        "source_row": request.row_number,
        "source_date": request.source_date,
        "comment_id": None,
        "parent_comment_id": None,
        "is_reply": None,
        "author_display_name": None,
        "author_channel_id": None,
        "comment_text": None,
        "like_count": None,
        "comment_published_at": None,
        "comment_updated_at": None,
        "reply_count": None,
    }
    if comment:
        row.update(
            {
                "comment_id": comment.comment_id,
                "parent_comment_id": comment.parent_comment_id,
                "is_reply": comment.is_reply,
                "author_display_name": comment.author_display_name,
                "author_channel_id": comment.author_channel_id,
                "comment_text": comment.text,
                "like_count": comment.like_count,
                "comment_published_at": comment.published_at,
                "comment_updated_at": comment.updated_at,
                "reply_count": comment.reply_count,
            }
        )
    return row


def _write_video_comments_csv(
    request: VideoRequest,
    metadata: YouTubeVideoMetadata,
    comments: list[YouTubeComment],
    *,
    download_root: Path,
    error: str | None = None,
) -> Path:
    output_dir = (
        download_root / _institution_folder(request.institution) / "comments"
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{request.video_id}_comments.csv"
    if error:
        rows = [_comment_row(request, metadata, status="failed", error=error)]
    elif comments:
        rows = [
            _comment_row(request, metadata, status="comment", comment=comment)
            for comment in comments
        ]
    else:
        rows = [_comment_row(request, metadata, status="no_comments")]

    with output_path.open("w", encoding="utf-8-sig", newline="") as target:
        writer = csv.DictWriter(target, fieldnames=COMMENT_FIELDS)
        writer.writeheader()
        writer.writerows(rows)
    return output_path


def _download_comments_one(
    request: VideoRequest,
    *,
    download_root: Path,
    youtube_client: YouTubeVideoApiClient,
    metadata: YouTubeVideoMetadata,
    include_replies: bool,
    max_comments: int | None,
) -> CommentDownloadResult:
    try:
        comments = youtube_client.fetch_comments(
            request.video_id,
            include_replies=include_replies,
            max_comments=max_comments,
        )
    except YouTubeApiError as exc:
        error = str(exc)
        output_path = _write_video_comments_csv(
            request,
            metadata,
            [],
            download_root=download_root,
            error=error,
        )
        return CommentDownloadResult(
            video_id=request.video_id,
            url=request.url,
            row_number=request.row_number,
            institution=request.institution,
            status="failed",
            comment_count=0,
            file_path=str(output_path),
            error=error,
        )

    output_path = _write_video_comments_csv(
        request,
        metadata,
        comments,
        download_root=download_root,
    )
    return CommentDownloadResult(
        video_id=request.video_id,
        url=request.url,
        row_number=request.row_number,
        institution=request.institution,
        status="completed" if comments else "no_comments",
        comment_count=len(comments),
        file_path=str(output_path),
    )


async def download_comments(
    requests: list[VideoRequest],
    *,
    download_root: Path,
    youtube_client: YouTubeVideoApiClient,
    workers: int = 1,
    include_replies: bool = True,
    max_comments: int | None = None,
) -> list[CommentDownloadResult]:
    if workers < 1:
        raise ValueError("workers debe ser mayor que cero.")
    if max_comments is not None and max_comments < 1:
        raise ValueError("max_comments debe ser mayor que cero.")

    download_root.mkdir(parents=True, exist_ok=True)
    video_ids = list(dict.fromkeys(request.video_id for request in requests))
    metadata_by_id = await asyncio.to_thread(
        youtube_client.fetch_video_metadata, video_ids
    )
    semaphore = asyncio.Semaphore(workers)
    progress_lock = asyncio.Lock()
    completed = 0

    async def run_one(request: VideoRequest) -> CommentDownloadResult:
        nonlocal completed
        async with semaphore:
            result = await asyncio.to_thread(
                _download_comments_one,
                request,
                download_root=download_root,
                youtube_client=youtube_client,
                metadata=metadata_by_id.get(request.video_id, YouTubeVideoMetadata()),
                include_replies=include_replies,
                max_comments=max_comments,
            )
        async with progress_lock:
            completed += 1
            print(f"[comments done {completed}/{len(requests)}] {request.video_id}")
        return result

    return await asyncio.gather(*(run_one(request) for request in requests))


def _write_report(report_dir: Path, results: list[DownloadResult]) -> None:
    report_dir.mkdir(parents=True, exist_ok=True)
    fields = ["video_id", "url", "row_number", "institution", "status", "file_path", "error"]
    with (report_dir / "downloads.csv").open("w", encoding="utf-8-sig", newline="") as target:
        writer = csv.DictWriter(target, fieldnames=fields)
        writer.writeheader()
        writer.writerows(asdict(result) for result in results)

    summary = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "requested": len(results),
        "completed": sum(result.status == "completed" for result in results),
        "failed": sum(result.status == "failed" for result in results),
    }
    (report_dir / "download_run.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def _write_comment_report(
    report_dir: Path, results: list[CommentDownloadResult]
) -> None:
    report_dir.mkdir(parents=True, exist_ok=True)
    fields = [
        "video_id",
        "url",
        "row_number",
        "institution",
        "status",
        "comment_count",
        "file_path",
        "error",
    ]
    with (report_dir / "comments_summary.csv").open(
        "w", encoding="utf-8-sig", newline=""
    ) as target:
        writer = csv.DictWriter(target, fieldnames=fields)
        writer.writeheader()
        writer.writerows(asdict(result) for result in results)

    summary = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "requested_videos": len(results),
        "completed": sum(result.status == "completed" for result in results),
        "without_comments": sum(
            result.status == "no_comments" for result in results
        ),
        "failed": sum(result.status == "failed" for result in results),
        "downloaded_comments": sum(result.comment_count for result in results),
    }
    (report_dir / "comment_download_run.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    _write_comments_workbook(report_dir, results)


def _write_comments_workbook(
    report_dir: Path, results: list[CommentDownloadResult]
) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError as exc:
        raise SearchDependencyError(
            "Falta openpyxl. Instala las dependencias con: pip install -r requirements.txt"
        ) from exc

    output_path = report_dir / "comentarios.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Comentarios"
    headers = ["Fecha", "Video", "Titulo", "Universidad", "Comentario", "Likes"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for result in results:
        source_path = Path(result.file_path)
        if not source_path.is_file():
            continue
        with source_path.open(encoding="utf-8-sig", newline="") as source:
            for row in csv.DictReader(source):
                if row.get("record_status") != "comment":
                    continue
                try:
                    likes = int(row.get("like_count") or 0)
                except ValueError:
                    likes = 0
                sheet.append(
                    [
                        row.get("comment_published_at"),
                        row.get("video_url"),
                        row.get("video_title"),
                        row.get("institution"),
                        row.get("comment_text"),
                        likes,
                    ]
                )
                video_cell = sheet.cell(row=sheet.max_row, column=2)
                if video_cell.value:
                    video_cell.hyperlink = str(video_cell.value)
                    video_cell.style = "Hyperlink"
                sheet.cell(row=sheet.max_row, column=5).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column, width in {
        "A": 23,
        "B": 45,
        "C": 45,
        "D": 38,
        "E": 80,
        "F": 10,
    }.items():
        sheet.column_dimensions[column].width = width
    workbook.save(output_path)
    workbook.close()
    return output_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Descarga videos o comentarios de YouTube desde un archivo XLSX."
    )
    source_group = parser.add_mutually_exclusive_group(required=True)
    source_group.add_argument("--input", type=Path, help="Ruta al archivo XLSX.")
    source_group.add_argument(
        "--video-id",
        action="append",
        help="ID o URL de un video. Puede repetirse para descargas puntuales.",
    )
    parser.add_argument("--sheet", help="Hoja del XLSX. Por defecto usa la hoja activa.")
    parser.add_argument(
        "--institution",
        help="Institucion obligatoria para --video-id; define su carpeta de descarga.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        help="Carpeta raiz de descarga. Por defecto: downloads/.",
    )
    parser.add_argument("--workers", type=int, default=1, help="Descargas simultaneas.")
    parser.add_argument("--retries", type=int, default=0, help="Reintentos por video.")
    parser.add_argument(
        "--format",
        dest="video_format",
        default="bv*[height<=720][ext=mp4]+ba[ext=m4a]/b[height<=720][ext=mp4]/b[ext=mp4]/b",
        help="Selector de formato de yt-dlp. Por defecto limita video a 720p.",
    )
    parser.add_argument("--cookies-from-browser", help="Por ejemplo: chrome, edge o firefox.")
    parser.add_argument("--cookies-browser-profile", help="Por ejemplo: Default o Profile 1.")
    parser.add_argument(
        "--download-comments",
        action="store_true",
        help="Descarga comentarios en CSV en lugar de descargar los videos.",
    )
    parser.add_argument(
        "--youtube-api-key",
        default=os.getenv("YOUTUBE_API_KEY"),
        help="API key de YouTube. Por defecto usa YOUTUBE_API_KEY.",
    )
    parser.add_argument(
        "--max-comments-per-video",
        type=int,
        help="Limite opcional de comentarios y respuestas por video.",
    )
    parser.add_argument(
        "--exclude-replies",
        action="store_true",
        help="Descarga solo comentarios principales, sin respuestas.",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    try:
        download_root = (
            args.output_dir.resolve()
            if args.output_dir
            else PROJECT_ROOT / "downloads"
        )
        if args.input:
            input_path = args.input.resolve()
            requests = load_video_requests(input_path, args.sheet)
            report_name = input_path.stem
        else:
            requests = build_direct_video_requests(args.video_id, args.institution or "")
            report_name = "manual"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if args.download_comments:
            if not args.youtube_api_key:
                raise ValueError(
                    "Falta la API key. Define YOUTUBE_API_KEY o usa --youtube-api-key."
                )
            print(
                f"Plan: comentarios de {len(requests)} videos, salida: {download_root}"
            )
            comment_results = asyncio.run(
                download_comments(
                    requests,
                    download_root=download_root,
                    youtube_client=YouTubeVideoApiClient(args.youtube_api_key),
                    workers=args.workers,
                    include_replies=not args.exclude_replies,
                    max_comments=args.max_comments_per_video,
                )
            )
            report_dir = (
                download_root
                / "_reports"
                / f"{report_name}_comments_{timestamp}"
            )
            _write_comment_report(report_dir, comment_results)
        else:
            print(f"Plan: {len(requests)} videos, salida: {download_root}")
            download_results = asyncio.run(
                download_videos(
                    requests,
                    download_root=download_root,
                    workers=args.workers,
                    video_format=args.video_format,
                    retries=args.retries,
                    cookies_from_browser=args.cookies_from_browser,
                    cookies_browser_profile=args.cookies_browser_profile,
                )
            )
            report_dir = download_root / "_reports" / f"{report_name}_{timestamp}"
            _write_report(report_dir, download_results)
    except (OSError, ValueError, SearchDependencyError, YouTubeApiError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    if args.download_comments:
        failed = sum(result.status == "failed" for result in comment_results)
        downloaded = sum(result.comment_count for result in comment_results)
        report_path = report_dir / "comentarios.xlsx"
        print(
            f"Comentarios terminados: {downloaded} registros de "
            f"{len(comment_results)} videos. Reporte: {report_path}"
        )
        return 0 if failed == 0 else 2

    completed = sum(result.status == "completed" for result in download_results)
    print(
        f"Descarga terminada: {completed}/{len(download_results)} videos. "
        f"Reporte: {report_dir / 'downloads.csv'}"
    )
    return 0 if completed == len(download_results) else 2


if __name__ == "__main__":
    raise SystemExit(main())
